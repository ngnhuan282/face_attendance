from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
import csv
import io

from accounts.constants import ADMIN_GROUP_NAME, TEACHER_GROUP_NAME
from accounts.permissions import group_required
from academics.models import Department

from .models import Student, StudentClass
from .forms import StudentForm, StudentClassForm


# ──────────────────────────────────────────────
# SINH VIÊN – CRUD
# ──────────────────────────────────────────────

@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def student_list(request):
    """Danh sách sinh viên – có tìm kiếm và lọc theo lớp / ngành."""
    qs = Student.objects.select_related('student_class', 'student_class__department')

    # Tìm kiếm
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(student_id__icontains=q) |
            Q(full_name__icontains=q) |
            Q(email__icontains=q)
        )

    # Lọc theo lớp
    class_id = request.GET.get('class_id', '').strip()
    if class_id:
        qs = qs.filter(student_class_id=class_id)

    # Lọc theo ngành
    dept_id = request.GET.get('dept_id', '').strip()
    if dept_id:
        qs = qs.filter(student_class__department_id=dept_id)

    # Thống kê
    total = Student.objects.count()
    with_photo = Student.objects.exclude(photo='').exclude(photo__isnull=True).count()
    without_photo = total - with_photo
    class_count = StudentClass.objects.count()

    # Dữ liệu cho filter dropdowns
    all_classes = StudentClass.objects.select_related('department').order_by('class_code')
    all_departments = Department.objects.all().order_by('name')

    # Phân trang – 10 dòng / trang
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'students/list.html', {
        'active_menu': 'students',
        'students': page_obj,          # page_obj thay cho qs
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'class_id': class_id,
        'dept_id': dept_id,
        'all_classes': all_classes,
        'all_departments': all_departments,
        # Stats
        'total': total,
        'with_photo': with_photo,
        'without_photo': without_photo,
        'class_count': class_count,
    })


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def student_create(request):
    """Thêm sinh viên mới (có upload ảnh khuôn mặt)."""
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            messages.success(request, f'Đã thêm sinh viên {student.full_name} ({student.student_id}) thành công.')
            return redirect('students:list')
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin. Có lỗi trong form.')
    else:
        form = StudentForm()

    return render(request, 'students/form.html', {
        'active_menu': 'students',
        'form': form,
        'form_title': 'Thêm Sinh Viên Mới',
        'submit_label': 'Thêm Sinh Viên',
        'cancel_url': 'students:list',
    })


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def student_edit(request, pk):
    """Sửa thông tin sinh viên."""
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật thông tin sinh viên {student.full_name}.')
            return redirect('students:list')
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = StudentForm(instance=student)

    return render(request, 'students/form.html', {
        'active_menu': 'students',
        'form': form,
        'student': student,
        'form_title': f'Sửa Sinh Viên — {student.full_name}',
        'submit_label': 'Lưu Thay Đổi',
        'cancel_url': 'students:list',
    })


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def student_delete(request, pk):
    """Xóa sinh viên (chỉ Admin) – chỉ chấp nhận POST qua AJAX."""
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        name = student.full_name
        sid = student.student_id
        student.delete()
        messages.success(request, f'Xóa sinh viên {name} ({sid}) thành công.')
        return JsonResponse({'success': True, 'message': f'Đã xóa sinh viên {name} ({sid}).'})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ──────────────────────────────────────────────
# LỚP SINH HOẠT – CRUD
# ──────────────────────────────────────────────

@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def studentclass_list(request):
    """Danh sách lớp sinh hoạt."""
    qs = StudentClass.objects.select_related('department').annotate(
        student_count=Count('students')
    ).order_by('-intake_year', 'class_code')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(class_code__icontains=q) | Q(class_name__icontains=q)
        )

    # Phân trang – 10 dòng / trang
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'students/class_list.html', {
        'active_menu': 'students',
        'classes': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
    })


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def studentclass_create(request):
    """Thêm lớp sinh hoạt."""
    if request.method == 'POST':
        form = StudentClassForm(request.POST)
        if form.is_valid():
            sc = form.save()
            messages.success(request, f'Đã tạo lớp {sc.class_code} thành công.')
            return redirect('students:class_list')
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = StudentClassForm()

    return render(request, 'students/form.html', {
        'active_menu': 'students',
        'form': form,
        'form_title': 'Thêm Lớp Sinh Hoạt',
        'submit_label': 'Tạo Lớp',
        'cancel_url': 'students:class_list',
    })


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def studentclass_edit(request, pk):
    """Sửa lớp sinh hoạt."""
    sc = get_object_or_404(StudentClass, pk=pk)

    if request.method == 'POST':
        form = StudentClassForm(request.POST, instance=sc)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật lớp {sc.class_code}.')
            return redirect('students:class_list')
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = StudentClassForm(instance=sc)

    return render(request, 'students/form.html', {
        'active_menu': 'students',
        'form': form,
        'sc': sc,
        'form_title': f'Sửa Lớp — {sc.class_code}',
        'submit_label': 'Lưu Thay Đổi',
        'cancel_url': 'students:class_list',
    })


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def studentclass_delete(request, pk):
    """Xóa lớp sinh hoạt (chỉ Admin, không xóa được nếu còn sinh viên) – chỉ chấp nhận POST qua AJAX."""
    sc = get_object_or_404(StudentClass, pk=pk)

    if request.method == 'POST':
        if sc.students.exists():
            return JsonResponse({
                'error': f'Không thể xóa lớp {sc.class_code} vì còn {sc.students.count()} sinh viên trong lớp.'
            }, status=400)
        code = sc.class_code
        sc.delete()
        messages.success(request, f'Xóa lớp sinh hoạt "{code}" thành công.')
        return JsonResponse({'success': True, 'message': f'Đã xóa lớp {code}.'})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def studentclass_detail(request, pk):
    """Xem danh sách sinh viên trong một lớp sinh hoạt."""
    sc = get_object_or_404(
        StudentClass.objects.select_related('department'),
        pk=pk
    )
    students_qs = sc.students.order_by('full_name')

    # Thêm chức năng tìm kiếm
    q = request.GET.get('q', '').strip()
    if q:
        students_qs = students_qs.filter(
            Q(student_id__icontains=q) |
            Q(full_name__icontains=q) |
            Q(email__icontains=q)
        )

    # Lấy tất cả sinh viên không thuộc lớp này để hiển thị trong phần thêm sinh viên
    available_students = Student.objects.exclude(student_class=sc).select_related('student_class').order_by('student_id')

    # Phân trang – 10 dòng / trang
    paginator = Paginator(students_qs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'students/class_detail.html', {
        'active_menu': 'students',
        'sc': sc,
        'students': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'available_students': available_students,
        'q': q,
    })


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def studentclass_add_student(request, pk):
    """Thêm sinh viên hiện có vào lớp sinh hoạt."""
    sc = get_object_or_404(StudentClass, pk=pk)
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        if student_id:
            try:
                student = Student.objects.select_related('student_class').get(pk=student_id)
                old_class = student.student_class
                student.student_class = sc
                student.save()
                if old_class:
                    messages.success(request, f'Đã chuyển sinh viên {student.full_name} ({student.student_id}) từ lớp {old_class.class_code} sang lớp {sc.class_code} thành công.')
                else:
                    messages.success(request, f'Đã thêm sinh viên {student.full_name} ({student.student_id}) vào lớp {sc.class_code} thành công.')
            except Student.DoesNotExist:
                messages.error(request, 'Không tìm thấy sinh viên đã chọn.')
        else:
            messages.error(request, 'Vui lòng chọn một sinh viên.')
    return redirect('students:class_detail', pk=pk)


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def studentclass_remove_student(request, pk):
    """Xóa sinh viên khỏi lớp sinh hoạt (set student_class = None) – chỉ chấp nhận POST qua AJAX."""
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        old_class = student.student_class
        student.student_class = None
        student.save()
        messages.success(request, f'Đã xóa sinh viên {student.full_name} ({student.student_id}) khỏi lớp {old_class.class_code if old_class else ""}.')
        return JsonResponse({'success': True, 'message': f'Đã xóa sinh viên {student.full_name} khỏi lớp.'})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ──────────────────────────────────────────────
# CHI TIẾT SINH VIÊN
# ──────────────────────────────────────────────

@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def student_detail(request, pk):
    """Trang chi tiết sinh viên: thông tin cá nhân + lịch sử điểm danh + tỉ lệ chuyên cần.

    Phân quyền:
    - Admin: xem mọi sinh viên.
    - GV: chỉ xem SV thuộc lớp HP mình dạy.
    """
    student = get_object_or_404(Student.objects.select_related('student_class', 'student_class__department'), pk=pk)

    # Kiểm tra quyền GV: chỉ được xem SV trong lớp HP mình dạy
    if request.is_teacher_group and not request.is_admin_group:
        from courses.models import Enrollment
        teacher = getattr(request.user, 'teacher', None)
        if teacher:
            # SV phải có enrollment trong ít nhất 1 lớp GV dạy
            is_allowed = Enrollment.objects.filter(
                student=student,
                course_class__teacher=teacher,
                is_active=True
            ).exists()
            if not is_allowed:
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied
        else:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

    # ── Lịch sử điểm danh theo từng lớp HP ──
    from attendance.models import AttendanceRecord, AttendanceSession
    from courses.models import Enrollment

    # Lấy tất cả enrollment của SV
    enrollments = (
        Enrollment.objects
        .filter(student=student)
        .select_related('course_class__course', 'course_class__semester', 'course_class__teacher__user')
    )

    attendance_by_class = []
    total_sessions_all = 0
    total_present_all = 0

    for enrollment in enrollments:
        cc = enrollment.course_class
        # Số buổi đã học trong lớp này
        sessions = AttendanceSession.objects.filter(course_class=cc)
        session_count = sessions.count()

        # Số buổi SV có mặt
        present_count = AttendanceRecord.objects.filter(
            session__course_class=cc,
            student=student,
            status='present'
        ).count()

        # Số buổi SV vắng
        absent_count = session_count - present_count

        attendance_rate = round((present_count / session_count * 100), 1) if session_count > 0 else None

        # Cảnh báo khi vắng > 20%
        is_warning = (attendance_rate is not None) and (attendance_rate < 80)

        # Lịch sử từng buổi của SV trong lớp này (10 buổi gần nhất)
        recent_records = (
            AttendanceRecord.objects
            .filter(session__course_class=cc, student=student)
            .select_related('session')
            .order_by('-session__started_at')[:10]
        )

        attendance_by_class.append({
            'course_class': cc,
            'enrollment': enrollment,
            'session_count': session_count,
            'present_count': present_count,
            'absent_count': absent_count,
            'attendance_rate': attendance_rate,
            'is_warning': is_warning,
            'recent_records': list(recent_records),
        })

        total_sessions_all += session_count
        total_present_all += present_count

    # ── Tỉ lệ chuyên cần tổng hợp ──
    overall_rate = round((total_present_all / total_sessions_all * 100), 1) if total_sessions_all > 0 else None
    overall_warning = (overall_rate is not None) and (overall_rate < 80)

    context = {
        'active_menu': 'students',
        'student': student,
        'attendance_by_class': attendance_by_class,
        'total_sessions_all': total_sessions_all,
        'total_present_all': total_present_all,
        'overall_rate': overall_rate,
        'overall_warning': overall_warning,
        'enrollment_count': enrollments.count(),
    }
    return render(request, 'students/detail.html', context)


# ──────────────────────────────────────────────
# IMPORT CSV SINH VIÊN HÀNG LOẠT
# ──────────────────────────────────────────────

@group_required(ADMIN_GROUP_NAME)
def student_import_csv(request):
    """Import hàng loạt sinh viên từ file CSV.

    Định dạng CSV:
        student_id, full_name, date_of_birth, email, phone, class_code

    Kết quả trả về JSON với:
        - success: số SV tạo thành công
        - errors: danh sách lỗi theo dòng
        - skipped: số dòng bỏ qua (MSSV đã tồn tại)

    GET: Trang form upload
    POST: Xử lý file CSV
    """
    if request.method == 'GET':
        all_classes = StudentClass.objects.select_related('department').order_by('class_code')
        return render(request, 'students/import_csv.html', {
            'active_menu': 'students',
            'all_classes': all_classes,
        })

    # ── POST: xử lý file CSV ──
    csv_file = request.FILES.get('csv_file')
    if not csv_file:
        return JsonResponse({'error': 'Chưa chọn file CSV.'}, status=400)

    if not csv_file.name.lower().endswith('.csv'):
        return JsonResponse({'error': 'File phải có định dạng .csv'}, status=400)

    if csv_file.size > 5 * 1024 * 1024:  # 5MB limit
        return JsonResponse({'error': 'File quá lớn (tối đa 5MB).'}, status=400)

    # Đọc nội dung — thử UTF-8-sig (BOM), rồi UTF-8, rồi latin-1
    try:
        raw = csv_file.read()
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                content = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return JsonResponse({'error': 'Không thể đọc file. Vui lòng lưu file với định dạng UTF-8.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Lỗi đọc file: {str(e)}'}, status=400)

    reader = csv.DictReader(io.StringIO(content))

    # Kiểm tra header
    required_fields = {'student_id', 'full_name'}
    if not reader.fieldnames:
        return JsonResponse({'error': 'File CSV trống hoặc không có header.'}, status=400)

    headers = {f.strip().lower() for f in reader.fieldnames}
    missing = required_fields - headers
    if missing:
        return JsonResponse({
            'error': f'File CSV thiếu cột bắt buộc: {", ".join(sorted(missing))}. '
                     f'Các cột hiện có: {", ".join(sorted(headers))}'
        }, status=400)

    created_count = 0
    skipped_count = 0
    errors = []

    try:
        with transaction.atomic():
            for row_num, row in enumerate(reader, start=2):
                # Chuẩn hóa key (bỏ khoảng trắng)
                row = {k.strip().lower(): (v.strip() if v else '') for k, v in row.items()}

                student_id = row.get('student_id', '').strip()
                full_name  = row.get('full_name', '').strip()

                # Validate bắt buộc
                if not student_id:
                    errors.append({'row': row_num, 'msg': 'Cột student_id trống.'})
                    continue
                if not full_name:
                    errors.append({'row': row_num, 'msg': f'{student_id}: Cột full_name trống.'})
                    continue

                # Bỏ qua nếu MSSV đã tồn tại
                if Student.objects.filter(student_id=student_id).exists():
                    skipped_count += 1
                    errors.append({'row': row_num, 'msg': f'{student_id}: MSSV đã tồn tại — bỏ qua.', 'type': 'skip'})
                    continue

                # Lớp sinh hoạt (tùy chọn)
                student_class = None
                class_code = row.get('class_code', '').strip()
                if class_code:
                    try:
                        student_class = StudentClass.objects.get(class_code=class_code)
                    except StudentClass.DoesNotExist:
                        errors.append({'row': row_num, 'msg': f'{student_id}: Không tìm thấy lớp "{class_code}" — SV sẽ không có lớp.'})
                    except StudentClass.MultipleObjectsReturned:
                        student_class = StudentClass.objects.filter(class_code=class_code).first()

                # Ngày sinh (tùy chọn) — chấp nhận dd/mm/yyyy hoặc yyyy-mm-dd
                dob = None
                dob_str = row.get('date_of_birth', '').strip()
                if dob_str:
                    from datetime import datetime
                    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                        try:
                            dob = datetime.strptime(dob_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    if dob is None:
                        errors.append({'row': row_num, 'msg': f'{student_id}: Ngày sinh "{dob_str}" không hợp lệ — bỏ qua trường này.'})

                # Email
                email = row.get('email', '').strip()
                phone = row.get('phone', '').strip()

                try:
                    with transaction.atomic():
                        Student.objects.create(
                            student_id=student_id,
                            full_name=full_name,
                            student_class=student_class,
                            date_of_birth=dob,
                            email=email,
                            phone=phone,
                            is_active=True,
                        )
                    created_count += 1
                except Exception as e:
                    errors.append({'row': row_num, 'msg': f'{student_id}: Lỗi tạo bản ghi — {str(e)}'})

    except Exception as e:
        return JsonResponse({'error': f'Lỗi xử lý file: {str(e)}'}, status=500)

    # Phân loại errors vs skips cho frontend
    real_errors = [e for e in errors if e.get('type') != 'skip']
    skips       = [e for e in errors if e.get('type') == 'skip']

    return JsonResponse({
        'success': True,
        'created': created_count,
        'skipped': skipped_count,
        'error_count': len(real_errors),
        'errors': real_errors,
        'skips': skips,
        'message': (
            f'Tạo thành công {created_count} sinh viên'
            + (f', bỏ qua {skipped_count} MSSV đã tồn tại' if skipped_count else '')
            + (f', có {len(real_errors)} lỗi' if real_errors else '')
            + '.'
        ),
    })


@group_required(ADMIN_GROUP_NAME)
def student_import_csv_template(request):
    """Tải xuống file CSV mẫu để hướng dẫn import."""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="mau_import_sinhvien.csv"'
    response.write('\ufeff')  # BOM for Excel

    writer = csv.writer(response)
    writer.writerow(['student_id', 'full_name', 'date_of_birth', 'email', 'phone', 'class_code'])
    writer.writerow(['SV220001', 'Nguyễn Văn A', '01/01/2004', 'sva@email.com', '0901234567', 'DHKTPM17A'])
    writer.writerow(['SV220002', 'Trần Thị B', '15/03/2004', 'svb@email.com', '0902345678', 'DHKTPM17A'])
    writer.writerow(['SV220003', 'Lê Văn C', '', '', '', ''])
    return response


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def student_export_excel(request):
    """Xuất danh sách sinh viên ra file Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from django.http import HttpResponse

    # Khởi tạo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DanhSachSinhVien"

    # Header
    headers = ['STT', 'MSSV', 'Họ Tên', 'Ngày Sinh', 'Email', 'Điện Thoại', 'Lớp Sinh Hoạt', 'Ngành']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Lấy dữ liệu giống như bộ lọc ở student_list
    qs = Student.objects.select_related('student_class', 'student_class__department')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(student_id__icontains=q) |
            Q(full_name__icontains=q) |
            Q(email__icontains=q)
        )

    class_id = request.GET.get('class_id', '').strip()
    if class_id:
        qs = qs.filter(student_class_id=class_id)

    dept_id = request.GET.get('dept_id', '').strip()
    if dept_id:
        qs = qs.filter(student_class__department_id=dept_id)

    qs = qs.order_by('student_class__class_code', 'student_id')

    # Ghi dữ liệu
    for row_num, student in enumerate(qs, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=student.student_id)
        ws.cell(row=row_num, column=3, value=student.full_name)
        ws.cell(row=row_num, column=4, value=student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '')
        ws.cell(row=row_num, column=5, value=student.email)
        ws.cell(row=row_num, column=6, value=student.phone)
        ws.cell(row=row_num, column=7, value=student.student_class.class_code if student.student_class else '')
        ws.cell(row=row_num, column=8, value=student.student_class.department.name if student.student_class and student.student_class.department else '')

    # Điều chỉnh độ rộng cột
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=danh_sach_sinh_vien.xlsx'
    wb.save(response)

    return response


# ──────────────────────────────────────────────
# HỒ SƠ SINH VIÊN (STUDENT SELF-SERVICE)
# ──────────────────────────────────────────────

from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash


def _student_required(view_func):
    """Decorator: yêu cầu đăng nhập và phải có student profile."""
    from functools import wraps
    from django.core.exceptions import PermissionDenied

    @login_required
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        student = getattr(request.user, 'student', None)
        if student is None:
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return wrapper


@_student_required
def student_profile(request):
    """Trang hồ sƠ cá nh​n cầu sinh viên.

    3 action (POST):
      - update_info : cập nhật email, phone
      - update_photo: đổi ảnh khuôn mật → re-encode
      - change_password: đổi mật khẩu
    """
    from .forms import StudentInfoForm, StudentPhotoForm, StudentPasswordForm
    from recognition.encoding_tasks import enqueue_student_encoding_update
    from courses.models import Enrollment
    from academics.models import Semester

    user = request.user
    student = user.student

    # --- Lấy học kỳ hiện tại (is_active=True) ---
    current_semester = Semester.objects.filter(is_active=True).first()
    enrollments = []
    if current_semester:
        enrollments = (
            Enrollment.objects
            .filter(student=student, course_class__semester=current_semester, is_active=True)
            .select_related(
                'course_class__course',
                'course_class__semester',
                'course_class__teacher__user',
                'course_class__course__room',
            )
            .order_by('course_class__course__course_code')
        )

    # Khởi tạo các form
    info_form = StudentInfoForm(initial={'email': student.email, 'phone': student.phone})
    photo_form = StudentPhotoForm()
    password_form = StudentPasswordForm(user_instance=user)

    if request.method == 'POST':
        action = request.POST.get('action', '')

        # ---- Cập nhật thông tin liên lạc ----
        if action == 'update_info':
            info_form = StudentInfoForm(request.POST)
            if info_form.is_valid():
                student.email = info_form.cleaned_data['email']
                student.phone = info_form.cleaned_data.get('phone', '')
                student.save(update_fields=['email', 'phone'])
                messages.success(request, 'Cập nhật thông tin thành công!')
                return redirect('students:student_profile')

        # ---- Cập nhật ảnh khuôn mật ----
        elif action == 'update_photo':
            photo_form = StudentPhotoForm(request.POST, request.FILES)
            if photo_form.is_valid():
                student.photo = photo_form.cleaned_data['photo']
                student.save(update_fields=['photo'])
                # Re-encode async
                enqueue_student_encoding_update(student.pk)
                messages.success(request, 'Đã cập nhật ảnh khuôn mật! Hệ thống đang xử lý nhận diện...')
                return redirect('students:student_profile')

        # ---- Đổi mật khẩu ----
        elif action == 'change_password':
            password_form = StudentPasswordForm(request.POST, user_instance=user)
            if password_form.is_valid():
                user.set_password(password_form.cleaned_data['new_password'])
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Đổi mật khẩu thành công!')
                return redirect('students:student_profile')

    context = {
        'student': student,
        'user': user,
        'info_form': info_form,
        'photo_form': photo_form,
        'password_form': password_form,
        'enrollments': enrollments,
        'current_semester': current_semester,
        'active_menu': 'profile',
    }
    return render(request, 'students/student_profile.html', context)

@_student_required
def student_timetable(request):
    """Giao diện Thời Khóa Biểu dành riêng cho sinh viên"""
    from academics.models import Semester
    from schedules.models import Schedule
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    student = request.user.student
    semesters = Semester.objects.all().order_by('-academic_year__start_date', '-semester_num')
    selected_semester = None
    
    semester_id = request.GET.get('semester')
    if semester_id:
        try:
            selected_semester = Semester.objects.get(id=semester_id)
        except Semester.DoesNotExist:
            pass
            
    if not selected_semester:
        selected_semester = Semester.objects.filter(is_active=True).first() or semesters.first()

    weeks = []
    today = timezone.now().date()
    target_date = today
    
    if selected_semester:
        sem_start_monday = selected_semester.start_date - timedelta(days=selected_semester.start_date.weekday())
        sem_end_sunday = selected_semester.end_date + timedelta(days=6 - selected_semester.end_date.weekday())
        
        total_weeks = (sem_end_sunday - sem_start_monday).days // 7 + 1
        for w in range(total_weeks):
            w_monday = sem_start_monday + timedelta(weeks=w)
            w_sunday = w_monday + timedelta(days=6)
            weeks.append({
                'number': w + 1,
                'monday': w_monday,
                'sunday': w_sunday,
                'label': f"Tuần {w + 1} ({w_monday.strftime('%d/%m/%Y')} - {w_sunday.strftime('%d/%m/%Y')})"
            })
            
        week_num = request.GET.get('week')
        if week_num:
            try:
                week_idx = int(week_num) - 1
                if 0 <= week_idx < len(weeks):
                    target_date = weeks[week_idx]['monday']
            except ValueError:
                pass
        else:
            current_week = next((w for w in weeks if w['monday'] <= today <= w['sunday']), None)
            if current_week:
                target_date = current_week['monday']
            else:
                target_date = weeks[0]['monday'] if weeks else today
                
    monday = target_date - timedelta(days=target_date.weekday())
    sunday = monday + timedelta(days=6)
    
    selected_week = next((w for w in weeks if w['monday'] == monday), None)
    
    week_days = []
    for i in range(7):
        day = monday + timedelta(days=i)
        week_days.append({
            'date': day,
            'day_name': f"Thứ {i+2}" if i < 6 else "Chủ Nhật",
            'is_today': day == today
        })
        
    schedules = Schedule.objects.select_related(
        'course_class', 'course_class__course', 'room', 'course_class__teacher__user'
    ).filter(
        date__range=[monday, sunday],
        course_class__enrollments__student=student,
        course_class__enrollments__is_active=True
    )
    
    grid = [[None for _ in range(7)] for _ in range(10)]
    for schedule in schedules:
        day_idx = (schedule.date - monday).days
        start_p = schedule.start_period - 1
        end_p = schedule.end_period - 1
        
        if 0 <= day_idx < 7 and 0 <= start_p < 10:
            schedule.rowspan = schedule.end_period - schedule.start_period + 1
            grid[start_p][day_idx] = schedule
            # Fill the spanned cells so the template can skip them
            for p in range(start_p + 1, min(end_p + 1, 10)):
                grid[p][day_idx] = 'spanned'

    context = {
        'active_menu': 'timetable',
        'semesters': semesters,
        'selected_semester': selected_semester,
        'weeks': weeks,
        'selected_week': selected_week,
        'target_date': target_date.strftime('%Y-%m-%d'),
        'monday': monday,
        'sunday': sunday,
        'week_days': week_days,
        'grid': grid,
        'periods': range(1, 11),
        'today': timezone.now().date(),
    }
    return render(request, 'students/timetable.html', context)

@_student_required
def student_attendance_history(request):
    """Giao diện lịch sử và thống kê điểm danh dành riêng cho sinh viên."""
    from attendance.models import AttendanceRecord, AttendanceSession
    from courses.models import Enrollment, CourseClass
    from academics.models import Semester
    
    student = request.user.student

    semesters = (
        Semester.objects
        .filter(course_classes__enrollments__student=student, course_classes__enrollments__is_active=True)
        .distinct()
        .order_by('-academic_year__start_date', '-semester_num')
    )

    selected_semester = None
    semester_id = request.GET.get('semester', '').strip()
    if semester_id:
        selected_semester = get_object_or_404(semesters, pk=semester_id)
    else:
        selected_semester = semesters.filter(is_active=True).first() or semesters.first()

    enrollments = (
        Enrollment.objects
        .filter(student=student, is_active=True)
        .select_related('course_class__course', 'course_class__semester', 'course_class__teacher__user')
        .order_by('-course_class__semester__start_date')
    )
    if selected_semester:
        enrollments = enrollments.filter(course_class__semester=selected_semester)

    filter_enrollments = enrollments
    display_enrollments = enrollments
    selected_class = None
    class_id = request.GET.get('class')
    if class_id:
        selected_class = get_object_or_404(
            CourseClass.objects.filter(
                enrollments__student=student,
                enrollments__is_active=True,
                enrollments__in=filter_enrollments,
            ),
            pk=class_id,
        )
        display_enrollments = display_enrollments.filter(course_class=selected_class)

    course_classes = [enrollment.course_class for enrollment in display_enrollments]
    course_class_ids = [course_class.id for course_class in course_classes]

    summary_cards = []
    total_sessions_all = 0
    total_present_all = 0
    total_late_all = 0
    total_effective_present_all = 0
    total_absent_all = 0

    for enrollment in display_enrollments:
        course_class = enrollment.course_class
        closed_sessions = AttendanceSession.objects.filter(
            course_class=course_class,
            status='closed',
            started_at__date__gte=enrollment.enrolled_at.date(),
        )
        total_sessions = closed_sessions.count()
        records = AttendanceRecord.objects.filter(
            session__in=closed_sessions,
            student=student,
        )
        present_count = records.filter(status='present').count()
        late_count = records.filter(status='late').count()
        effective_present = present_count + late_count
        absent_count = max(total_sessions - effective_present, 0)
        attendance_rate = round(effective_present / total_sessions * 100, 1) if total_sessions else 0.0
        absent_rate = round(absent_count / total_sessions * 100, 1) if total_sessions else 0.0

        summary_cards.append({
            'course_class': course_class,
            'total_sessions': total_sessions,
            'present_count': present_count,
            'late_count': late_count,
            'absent_count': absent_count,
            'attendance_rate': attendance_rate,
            'absent_rate': absent_rate,
            'is_danger': absent_rate > 20,
        })

        total_sessions_all += total_sessions
        total_present_all += present_count
        total_late_all += late_count
        total_effective_present_all += effective_present
        total_absent_all += absent_count

    overall_attendance_rate = round(total_effective_present_all / total_sessions_all * 100, 1) if total_sessions_all else 0.0
    overall_absent_rate = round(total_absent_all / total_sessions_all * 100, 1) if total_sessions_all else 0.0
    has_exam_risk = any(card['is_danger'] for card in summary_cards)

    sessions_qs = (
        AttendanceSession.objects
        .filter(course_class_id__in=course_class_ids)
        .select_related(
            'course_class__course',
            'course_class__semester',
            'schedule',
            'schedule__room',
        )
        .order_by('-started_at')
    )
        
    records = AttendanceRecord.objects.filter(student=student, session__in=sessions_qs)
    record_map = {r.session_id: r for r in records}

    history = []
    for s in sessions_qs:
        rec = record_map.get(s.id)
        if rec:
            status, method, conf, note = rec.status, rec.method, rec.confidence, rec.note
        else:
            status = 'absent' if s.status == 'closed' else 'pending'
            method, conf, note = '', 0.0, ''

        schedule = s.schedule
            
        history.append({
            'session': s,
            'date': schedule.date if schedule else s.started_at.date(),
            'period_text': f"{schedule.start_period} - {schedule.end_period}" if schedule else "—",
            'room': schedule.room if schedule else None,
            'session_number': schedule.session_number if schedule else None,
            'status': status,
            'method': method,
            'confidence': conf,
            'note': note,
        })

    total_pending_all = sum(1 for item in history if item['status'] == 'pending')

    from django.core.paginator import Paginator
    history_paginator = Paginator(history, 10)
    history_page_obj = history_paginator.get_page(request.GET.get('page', 1))

    return render(request, 'students/attendance_history.html', {
        'student': student,
        'semesters': semesters,
        'selected_semester': selected_semester,
        'enrollments': filter_enrollments,
        'selected_class': selected_class,
        'summary_cards': summary_cards,
        'overall_attendance_rate': overall_attendance_rate,
        'overall_absent_rate': overall_absent_rate,
        'total_sessions_all': total_sessions_all,
        'total_present_all': total_present_all,
        'total_late_all': total_late_all,
        'total_absent_all': total_absent_all,
        'total_pending_all': total_pending_all,
        'has_exam_risk': has_exam_risk,
        'history_page_obj': history_page_obj,
        'active_menu': 'attendance_history',
    })
