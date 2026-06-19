from __future__ import annotations

import io
from datetime import datetime

from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from accounts.constants import ADMIN_GROUP_NAME, TEACHER_GROUP_NAME
from accounts.permissions import group_required
from academics.models import Semester
from attendance.models import AttendanceRecord, AttendanceSession
from courses.models import CourseClass, Enrollment
from students.models import Student

from .models import AttendanceReport
from .services import refresh_class_reports



def _build_session_columns(course_class: CourseClass):
    return list(
        AttendanceSession.objects.filter(
            course_class=course_class,
            status='closed',
        ).order_by('started_at')
    )


def _build_matrix(reports, sessions):
    session_ids = [s.pk for s in sessions]
    student_ids = [r.student_id for r in reports]

    # Lấy tất cả record
    records = AttendanceRecord.objects.filter(
        session_id__in=session_ids,
        student__in=student_ids,
    ).values('session_id', 'student_id', 'status')

    # Index: (session_id, student_id) → status
    rec_map = {(r['session_id'], r['student_id']): r['status'] for r in records}

    # Fetch enrollments to know when each student enrolled
    course_class_id = sessions[0].course_class_id if sessions else None
    enrollments = {}
    if course_class_id:
        enr_list = Enrollment.objects.filter(
            course_class_id=course_class_id,
            student_id__in=student_ids,
            is_active=True,
        ).values('student_id', 'enrolled_at')
        enrollments = {e['student_id']: e['enrolled_at'].date() for e in enr_list}

    rows = []
    for rpt in reports:
        enrolled_date = enrollments.get(rpt.student_id)
        cells = []
        for s in sessions:
            # Nếu buổi học diễn ra trước khi SV đăng ký, không tính vắng
            if enrolled_date and s.started_at.date() < enrolled_date:
                cells.append('not_enrolled')
            else:
                cells.append(rec_map.get((s.pk, rpt.student_id), 'absent'))
        rows.append({'report': rpt, 'cells': cells})
    return rows


def _get_active_reports(course_class: CourseClass):
    return (
        AttendanceReport.objects
        .filter(
            course_class=course_class,
            student__enrollments__course_class=course_class,
            student__enrollments__is_active=True,
        )
        .select_related('student__student_class__department')
        .order_by('student__full_name')
        .distinct()
    )


def _teacher_scope(request):
    if request.is_teacher_group and not request.is_admin_group:
        teacher = getattr(request.user, 'teacher', None)
        if teacher is None:
            raise PermissionDenied
        return teacher
    return None


# Trang chọn học kỳ + lớp HP
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def report_index(request):
    semesters    = Semester.objects.all().order_by('-start_date')
    selected_sem = None
    course_classes = CourseClass.objects.none()

    sem_id = request.GET.get('semester')
    page_obj = None
    if sem_id:
        selected_sem   = get_object_or_404(Semester, pk=sem_id)
        course_classes = (
            CourseClass.objects
            .filter(semester=selected_sem)
            .select_related('course', 'teacher__user')
            .order_by('class_code')
        )
        # GV chỉ được xem lớp mình dạy
        teacher = _teacher_scope(request)
        if teacher:
            course_classes = course_classes.filter(teacher=teacher)
        from django.core.paginator import Paginator
        paginator = Paginator(course_classes, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

    return render(request, 'reports/index.html', {
        'semesters'     : semesters,
        'selected_sem'  : selected_sem,
        'course_classes': course_classes,
        'page_obj'      : page_obj,
        'active_menu'   : 'reports',
    })


# Bảng chuyên cần + biểu đồ cột theo buổi
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def report_class(request, class_id):
    course_class = get_object_or_404(
        CourseClass.objects.select_related('course', 'semester', 'teacher__user'),
        pk=class_id,
    )

    # GV chỉ được xem báo cáo lớp mình dạy
    teacher = _teacher_scope(request)
    if teacher and course_class.teacher_id != teacher.pk:
        raise PermissionDenied

    if request.GET.get('refresh') == '1':
        refresh_class_reports(course_class)

    reports = _get_active_reports(course_class)

    sessions = _build_session_columns(course_class)
    matrix   = _build_matrix(reports, sessions)

    # Thống kê cho stat cards
    total_students = reports.count()
    good_count     = reports.filter(attendance_rate__gte=80).count()
    banned_count   = reports.filter(attendance_rate__lt=80).count()
    warning_count  = reports.filter(absent_rate__gte=20, absent_rate__lt=40).count()
    danger_count   = reports.filter(absent_rate__gte=40).count()
    avg_rate       = 0.0
    if total_students:
        avg_rate = round(sum(r.attendance_rate for r in reports) / total_students, 1)

    # Fetch active enrollments once for chart calculation
    enr_list = Enrollment.objects.filter(
        course_class=course_class,
        is_active=True
    ).values('student_id', 'enrolled_at')
    enrollment_dates = {e['student_id']: e['enrolled_at'].date() for e in enr_list}

    chart_labels  = [f"Buổi {i+1}" for i in range(len(sessions))]
    chart_present = []
    chart_absent  = []
    for s in sessions:
        # Chỉ đếm những sinh viên đã đăng ký trước hoặc trong ngày buổi học diễn ra
        eligible_student_ids = [
            student_id
            for student_id, enr_date in enrollment_dates.items()
            if enr_date <= s.started_at.date()
        ]
        present = AttendanceRecord.objects.filter(
            session=s,
            student_id__in=eligible_student_ids,
            status__in=['present', 'late'],
        ).count()
        absent  = len(eligible_student_ids) - present

        chart_present.append(present)
        chart_absent.append(absent)

    from django.core.paginator import Paginator
    paginator = Paginator(matrix, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'reports/class_report.html', {
        'course_class'  : course_class,
        'reports'       : reports,
        'sessions'      : sessions,
        'page_obj'      : page_obj,
        'total_students': total_students,
        'good_count'    : good_count,
        'banned_count'  : banned_count,
        'warning_count' : warning_count,
        'danger_count'  : danger_count,
        'avg_rate'      : avg_rate,
        'chart_labels'  : chart_labels,
        'chart_present' : chart_present,
        'chart_absent'  : chart_absent,
        'active_menu'   : 'reports',
    })



# Xuất Excel
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def export_excel(request, class_id):
    """Xuất báo cáo chuyên cần theo template danh sách sinh viên của nhóm."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    course_class = get_object_or_404(
        CourseClass.objects.select_related('course', 'semester', 'teacher__user'),
        pk=class_id,
    )

    # GV chỉ được xuất báo cáo lớp mình dạy
    teacher = _teacher_scope(request)
    if teacher and course_class.teacher_id != teacher.pk:
        raise PermissionDenied

    reports = _get_active_reports(course_class).order_by('student__student_id')
    sessions = _build_session_columns(course_class)
    matrix   = _build_matrix(reports, sessions)

    student_ids = [row['report'].student_id for row in matrix]
    enrollment_map = {
        enrollment.student_id: enrollment
        for enrollment in Enrollment.objects.filter(
            course_class=course_class,
            student_id__in=student_ids,
        ).select_related('student')
    }

    header_fill = PatternFill('solid', fgColor='1A6B3C')
    present_fill = PatternFill('solid', fgColor='DCFCE7')
    absent_fill = PatternFill('solid', fgColor='FEE2E2')
    late_fill = PatternFill('solid', fgColor='FEF9C3')
    excused_fill = PatternFill('solid', fgColor='E0E7FF')

    header_font = Font(bold=True, color='FFFFFF')
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(italic=True)
    present_font = Font(color='166534')
    absent_font = Font(color='991B1B')
    late_font = Font(color='854D0E')
    excused_font = Font(color='3730A3')

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Danh sách sinh viên'

    ws.merge_cells('A1:H1')
    title_cell = ws.cell(1, 1, f'DANH SÁCH SINH VIÊN - LỚP {course_class.class_code}')
    title_cell.font = title_font
    title_cell.alignment = align_center

    ws.merge_cells('A2:H2')
    teacher_name = course_class.teacher.user.get_full_name() or course_class.teacher.user.username
    subtitle_cell = ws.cell(
        2,
        1,
        f'Môn học: {course_class.course.course_name} | Giảng viên: {teacher_name}',
    )
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = align_center

    HEADER_ROW = 4
    headers = [
        'STT',
        'MSSV',
        'Họ Tên',
        'Lớp Sinh Hoạt',
        'Ngành',
        'Trạng Thái',
        'Ngày Đăng Ký',
        'Tỉ lệ đi học',
    ]
    headers.extend(
        session.started_at.strftime('%d/%m/%Y') if session.started_at else ''
        for session in sessions
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(HEADER_ROW, col_num, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    status_styles = {
        'present': ('Có mặt', present_fill, present_font),
        'absent': ('Vắng', absent_fill, absent_font),
        'late': ('Đi trễ', late_fill, late_font),
        'excused': ('Có phép', excused_fill, excused_font),
        'not_enrolled': ('-', None, None),
    }

    session_totals = [
        {'present': 0, 'absent': 0, 'late': 0}
        for _ in sessions
    ]

    for row_num, row_data in enumerate(matrix, 1):
        rpt = row_data['report']
        sv = rpt.student
        student_class = sv.student_class
        enrollment = enrollment_map.get(sv.pk)
        excel_row = HEADER_ROW + row_num

        row = [
            row_num,
            sv.student_id,
            sv.full_name,
            student_class.class_code if student_class else '',
            student_class.department.name if student_class and student_class.department else '',
            'Đang học' if enrollment and enrollment.is_active else 'Nghỉ học',
            enrollment.enrolled_at.strftime('%d/%m/%Y') if enrollment else '',
            f'{rpt.attendance_rate:.1f}%',
        ]

        for col_num, value in enumerate(row, 1):
            cell = ws.cell(excel_row, col_num, value)
            cell.border = border
            cell.alignment = align_left if col_num in (3, 5) else align_center

        for index, status in enumerate(row_data['cells'], start=9):
            label, fill, font = status_styles.get(status, ('-', None, None))
            cell = ws.cell(excel_row, index, label)
            cell.border = border
            cell.alignment = align_center
            if fill:
                cell.fill = fill
            if font:
                cell.font = font

            if index >= 9:
                session_index = index - 9
                if session_index < len(session_totals) and status in session_totals[session_index]:
                    session_totals[session_index][status] += 1

    total_row = HEADER_ROW + len(matrix) + 1
    summary_fill = PatternFill('solid', fgColor='E2EFDA')
    summary_font = Font(bold=True, color='166534')
    avg_rate = sum(row['report'].attendance_rate for row in matrix) / len(matrix) if matrix else 0.0
    summary_values = [
        'Tổng kết',
        '',
        f'{len(matrix)} sinh viên',
        '',
        '',
        '',
        '',
        f'{avg_rate:.1f}%',
    ]
    for col_num, value in enumerate(summary_values, 1):
        cell = ws.cell(total_row, col_num, value)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = border
        cell.alignment = align_left if col_num == 3 else align_center

    for index, totals in enumerate(session_totals, start=9):
        cell = ws.cell(
            total_row,
            index,
            f"Có mặt: {totals['present']}\nVắng: {totals['absent']}\nTrễ: {totals['late']}",
        )
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[total_row].height = 48

    column_widths = {
        'A': 6,
        'B': 15,
        'C': 30,
        'D': 15,
        'E': 25,
        'F': 15,
        'G': 15,
        'H': 12,
    }
    for idx in range(len(sessions)):
        column_widths[get_column_letter(idx + 9)] = 15
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    filename = f"bao_cao_chuyen_can_{course_class.class_code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



# Lịch sử điểm danh từng SV
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def student_attendance(request, student_id):
    """
    Trang chi tiết lịch sử điểm danh của 1 sinh viên.
    Lọc theo lớp HP nếu có ?class=<id>.
    """
    student = get_object_or_404(Student.objects.select_related('student_class'), pk=student_id)
    teacher = _teacher_scope(request)

    # Lấy tất cả lớp HP SV đã đăng ký
    enrollments = (
        Enrollment.objects
        .filter(student=student)
        .select_related('course_class__course', 'course_class__semester')
        .order_by('-course_class__semester__start_date')
    )

    # GV chỉ được xem SV trong lớp mình dạy
    if teacher:
        enrollments = enrollments.filter(course_class__teacher=teacher)
        if not enrollments.filter(is_active=True).exists():
            raise PermissionDenied

    selected_class = None
    class_id = request.GET.get('class')
    if class_id:
        course_class_qs = CourseClass.objects.all()
        if teacher:
            course_class_qs = course_class_qs.filter(teacher=teacher)
        selected_class = get_object_or_404(course_class_qs, pk=class_id)

    # Lấy các session của lớp SV có đăng ký
    sessions_qs = (
        AttendanceSession.objects
        .filter(course_class__enrollments__student=student, course_class__enrollments__is_active=True)
        .select_related('course_class__course', 'course_class__semester')
        .order_by('-started_at')
    )
    if teacher:
        sessions_qs = sessions_qs.filter(course_class__teacher=teacher)
    if selected_class:
        sessions_qs = sessions_qs.filter(course_class=selected_class)
        
    records = AttendanceRecord.objects.filter(student=student, session__in=sessions_qs)
    record_map = {r.session_id: r for r in records}

    history = []
    for s in sessions_qs:
        rec = record_map.get(s.id)
        if rec:
            status, method, conf, note = rec.status, rec.method, rec.confidence, rec.note
            rec_id = rec.pk
        else:
            status = 'absent' if s.status == 'closed' else 'pending'
            method, conf, note = '', 0.0, ''
            rec_id = None
            
        history.append({
            'session': s,
            'status': status,
            'method': method,
            'confidence': conf,
            'note': note,
            'rec_id': rec_id
        })

    from django.core.paginator import Paginator
    history_paginator = Paginator(history, 10)
    history_page_obj = history_paginator.get_page(request.GET.get('page', 1))

    # Tổng hợp theo lớp HP
    reports = (
        AttendanceReport.objects
        .filter(student=student)
        .select_related('course_class__course', 'course_class__semester')
        .order_by('-course_class__semester__start_date')
    )
    if teacher:
        reports = reports.filter(course_class__teacher=teacher)
    if selected_class:
        reports = reports.filter(course_class=selected_class)

    return render(request, 'reports/student_attendance.html', {
        'student'       : student,
        'enrollments'   : enrollments,
        'selected_class': selected_class,
        'history'       : history,
        'history_page_obj': history_page_obj,
        'reports'       : reports,
        'active_menu'   : 'reports',
    })


# Sửa thủ công 1 AttendanceRecord
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(['POST'])
def attendance_edit(request, session_id, student_id):
    """
    Cho phép GV sửa trạng thái điểm danh khi nhận diện sai.
    POST: status (present|absent|late), note
    Sau khi sửa → refresh báo cáo + kiểm tra cảnh báo.
    """
    from reports.services import refresh_report
    from notifications.services import check_and_notify

    session = get_object_or_404(
        AttendanceSession.objects.select_related('course_class__teacher'),
        pk=session_id,
    )
    student = get_object_or_404(Student, pk=student_id)
    teacher = _teacher_scope(request)

    if teacher and session.course_class.teacher_id != teacher.pk:
        raise PermissionDenied
    if not Enrollment.objects.filter(
        course_class=session.course_class,
        student=student,
        is_active=True,
    ).exists():
        raise PermissionDenied

    new_status = request.POST.get('status')
    if new_status not in ('present', 'absent', 'late'):
        messages.error(request, 'Trạng thái không hợp lệ.')
        return redirect(request.META.get('HTTP_REFERER', '/'))

    record, created = AttendanceRecord.objects.get_or_create(
        session=session,
        student=student,
        defaults={
            'status': new_status,
            'method': 'manual',
            'note': request.POST.get('note', '').strip()[:200]
        }
    )

    if not created:
        record.status = new_status
        record.method = 'manual'
        record.note   = request.POST.get('note', '').strip()[:200]
        record.save(update_fields=['status', 'method', 'note'])

    # Cập nhật báo cáo + kiểm tra cảnh báo
    report = refresh_report(student, session.course_class)
    check_and_notify(student, session.course_class, report)

    messages.success(
        request,
        f'Đã cập nhật điểm danh {student.full_name} → '
        f'{dict([("present","Có mặt"), ("late","Đi trễ"), ("absent","Vắng")]).get(new_status)}'
    )

    # Quay lại trang chi tiết SV, đúng lớp HP
    return redirect(
        f'/reports/student/{student.pk}/?class={session.course_class.pk}'
    )
