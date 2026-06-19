import json
import logging
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor

import cv2
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from accounts.constants import ADMIN_GROUP_NAME, TEACHER_GROUP_NAME
from accounts.permissions import group_required
from courses.models import CourseClass, Enrollment
from recognition.face_detector import FaceDetector, draw_faces, open_webcam
from recognition.face_matcher import recognize_faces_in_frame, recognize_from_image
from recognition.utils import load_encodings
from schedules.models import Schedule
from students.models import Student

from .models import AttendanceRecord, AttendanceSession


STREAM_WIDTH = 640
STREAM_JPEG_QUALITY = 65
DETECTION_INTERVAL_SECONDS = 0.1
RECOGNITION_INTERVAL_SECONDS = 1.5
logger = logging.getLogger(__name__)


def _json_body(request):
    if not request.body:
        return {}
    try:
        return json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise ValidationError(f"JSON khong hop le: {exc}")


def _ok(data=None, status=200):
    return JsonResponse(data or {}, status=status, json_dumps_params={"ensure_ascii": False})


def _error(message, status=400):
    return _ok({"error": message}, status=status)


def _session_payload(session):
    return {
        "id": session.id,
        "course_class_id": session.course_class_id,
        "course_class": session.course_class.class_code,
        "schedule_id": session.schedule_id,
        "created_by_id": session.created_by_id,
        "started_at": session.started_at.isoformat() if session.started_at else None,
        "ended_at": session.ended_at.isoformat() if session.ended_at else None,
        "status": session.status,
        "note": session.note,
    }


def _record_payload(record):
    return {
        "id": record.id,
        "session_id": record.session_id,
        "student_id": record.student_id,
        "student_code": record.student.student_id,
        "student_name": record.student.full_name,
        "status": record.status,
        "method": record.method,
        "confidence": record.confidence,
        "timestamp": record.timestamp.isoformat() if record.timestamp else None,
        "note": record.note,
    }


def _default_user(request):
    if request.user.is_authenticated:
        return request.user
    return User.objects.order_by("id").first()


def _is_teacher_only(request):
    return getattr(request, "is_teacher_group", False) and not getattr(request, "is_admin_group", False)


def _teacher_for_request(request):
    teacher = getattr(request.user, "teacher", None)
    if not teacher:
        raise PermissionDenied
    return teacher


def _scope_course_classes(request, queryset):
    if _is_teacher_only(request):
        return queryset.filter(teacher=_teacher_for_request(request))
    return queryset


def _scope_sessions(request, queryset):
    if _is_teacher_only(request):
        return queryset.filter(course_class__teacher=_teacher_for_request(request))
    return queryset


def _check_course_class_access(request, course_class):
    if _is_teacher_only(request) and course_class.teacher_id != _teacher_for_request(request).id:
        raise PermissionDenied
    return course_class


def _check_session_access(request, session):
    _check_course_class_access(request, session.course_class)
    return session


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
def attendance_demo(request):
    course_classes = _scope_course_classes(
        request,
        CourseClass.objects.select_related("course", "semester", "teacher").order_by(
            "semester", "class_code"
        ),
    )
    selected_session = None
    students = []
    records_by_student = {}
    student_rows = []
    error_message = ""

    if request.method == "POST":
        course_class_id = request.POST.get("course_class_id")
        schedule_id = request.POST.get("schedule_id")
        note = request.POST.get("note", "")

        try:
            if schedule_id:
                schedule = get_object_or_404(Schedule, pk=schedule_id)
                course_class = schedule.course_class
                _check_course_class_access(request, course_class)
                if hasattr(schedule, 'attendance_session') and schedule.attendance_session:
                    return redirect(f"{reverse('attendance:demo')}?session_id={schedule.attendance_session.id}")
                selected_session = AttendanceSession.objects.create(
                    course_class=course_class,
                    schedule=schedule,
                    created_by=request.user,
                    note=note,
                )
            else:
                course_class = get_object_or_404(CourseClass, pk=course_class_id)
                _check_course_class_access(request, course_class)
                selected_session = AttendanceSession.objects.create(
                    course_class=course_class,
                    created_by=request.user,
                    note=note,
                )
            return redirect(f"{reverse('attendance:demo')}?session_id={selected_session.id}")
        except (ValidationError, IntegrityError) as exc:
            error_message = str(exc)

    session_id = request.GET.get("session_id")
    if session_id:
        selected_session = get_object_or_404(
            _scope_sessions(
                request,
                AttendanceSession.objects.select_related(
                    "course_class",
                    "course_class__course",
                    "course_class__teacher",
                    "created_by",
                ),
            ),
            pk=session_id,
        )
        students = [
            enrollment.student
            for enrollment in Enrollment.objects.select_related("student", "student__student_class")
            .filter(course_class=selected_session.course_class, is_active=True)
            .order_by("student__student_id")
        ]
        records_by_student = {
            record.student_id: record
            for record in AttendanceRecord.objects.filter(session=selected_session).select_related("student")
        }
        student_rows = [
            {
                "student": student,
                "record": records_by_student.get(student.id),
            }
            for student in students
        ]
        paginator = Paginator(student_rows, 10)
        page_obj = paginator.get_page(request.GET.get("page"))
    else:
        page_obj = None

    return render(
        request,
        "attendance/create_session.html",
        {
            "active_menu": "attendance",
            "course_classes": course_classes,
            "selected_session": selected_session,
            "session_id_json": selected_session.id if selected_session else None,
            "students": students,
            "student_rows": student_rows,
            "page_obj": page_obj,
            "present_records_count": sum(
                1 for record in records_by_student.values() if record.status == "present"
            ),
            "error_message": error_message,
        },
    )


def _resize_stream_frame(frame):
    height, width = frame.shape[:2]
    if width <= STREAM_WIDTH:
        return frame

    target_height = int(height * STREAM_WIDTH / width)
    return cv2.resize(frame, (STREAM_WIDTH, target_height), interpolation=cv2.INTER_AREA)


def _record_face_attendance(session, student, confidence, note):
    """Mark a student present, but never override manual late/absent decisions."""
    with transaction.atomic():
        locked_session = AttendanceSession.objects.select_for_update().get(pk=session.pk)
        if locked_session.status != "open":
            return None, False, False

        record, created = AttendanceRecord.objects.select_for_update().get_or_create(
            session=locked_session,
            student=student,
            defaults={
                "status": "present",
                "method": "face",
                "confidence": confidence,
                "timestamp": timezone.now(),
                "note": note,
            },
        )
        if created or record.status == "present":
            return record, created, not created

        if record.method == "manual" and record.status in {"late", "absent"}:
            return record, False, False

        record.status = "present"
        record.method = "face"
        record.confidence = confidence
        record.timestamp = timezone.now()
        record.note = note
        record.save(update_fields=["status", "method", "confidence", "timestamp", "note"])
        return record, False, False


def _mark_unrecorded_students_absent(session):
    recorded_student_ids = AttendanceRecord.objects.filter(session=session).values_list("student_id", flat=True)
    missing_student_ids = (
        Enrollment.objects.filter(course_class=session.course_class, is_active=True)
        .exclude(student_id__in=recorded_student_ids)
        .values_list("student_id", flat=True)
    )
    now = timezone.now()
    absent_records = [
        AttendanceRecord(
            session=session,
            student_id=student_id,
            status="absent",
            method="manual",
            timestamp=now,
            note="Tự động đánh vắng khi kết thúc buổi điểm danh",
        )
        for student_id in missing_student_ids
    ]
    if absent_records:
        AttendanceRecord.objects.bulk_create(absent_records, ignore_conflicts=True)
    return len(absent_records)


def _camera_label(student):
    label = f"{student.student_id} - {student.full_name}"
    normalized = unicodedata.normalize("NFD", label)
    return "".join(character for character in normalized if not unicodedata.combining(character)).replace("Đ", "D").replace("đ", "d")


def _mark_recognized_students(session, enrolled_students, marked_students, matches):
    labels = []
    for match in matches:
        student = enrolled_students.get(match.get("student_id")) if match else None
        if not student:
            labels.append("Unknown")
            continue

        labels.append(_camera_label(student))
        if student.student_id not in marked_students:
            record, _created, _already_present = _record_face_attendance(
                session,
                student,
                match["confidence"],
                "Auto recognized from MJPEG stream",
            )
            if record:
                marked_students.add(student.student_id)

    return labels


def _stream_frames(webcam, session, data):
    overlay_detector = FaceDetector()
    recognition_detector = FaceDetector()
    enrolled_students = {
        student.student_id: student
        for student in Student.objects.filter(
            enrollments__course_class=session.course_class,
            enrollments__is_active=True,
            is_active=True,
        )
    }
    marked_students = set(
        AttendanceRecord.objects.filter(session=session, status="present").values_list(
            "student__student_id", flat=True
        )
    )
    face_locations = []
    labels = []
    recognition_future = None
    recognition_started_at = 0.0
    detection_started_at = 0.0
    recognition_worker = ThreadPoolExecutor(max_workers=1, thread_name_prefix="attendance-recognition")

    try:
        while True:
            ok, frame = webcam.read()
            if not ok:
                break

            frame = _resize_stream_frame(frame)
            now = time.monotonic()
            if now - detection_started_at >= DETECTION_INTERVAL_SECONDS:
                latest_locations = overlay_detector.detect_faces(frame)
                if len(latest_locations) != len(face_locations):
                    labels = ["Unknown"] * len(latest_locations)
                face_locations = latest_locations
                detection_started_at = now

            if recognition_future and recognition_future.done():
                try:
                    _recognized_locations, matches = recognition_future.result()
                    labels = _mark_recognized_students(session, enrolled_students, marked_students, matches)
                except Exception:
                    logger.exception("Loi xu ly frame nhan dien cho session %s", session.pk)
                    labels = ["Unknown"] * len(face_locations)
                recognition_future = None

            if (
                recognition_future is None
                and now - recognition_started_at >= RECOGNITION_INTERVAL_SECONDS
            ):
                recognition_future = recognition_worker.submit(
                    recognize_faces_in_frame,
                    frame.copy(),
                    data,
                    0.5,
                    recognition_detector,
                )
                recognition_started_at = now

            draw_faces(frame, face_locations, labels)
            encoded, buffer = cv2.imencode(
                ".jpg",
                frame,
                [cv2.IMWRITE_JPEG_QUALITY, STREAM_JPEG_QUALITY],
            )
            if not encoded:
                continue

            yield (
                b"--frame\r\n"
                b"Content-Type: image/jpeg\r\n\r\n"
                + buffer.tobytes()
                + b"\r\n"
            )
    finally:
        recognition_worker.shutdown(wait=False, cancel_futures=True)
        webcam.release()


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(["GET"])
def video_stream(request, session_id):
    session = get_object_or_404(
        _scope_sessions(
            request,
            AttendanceSession.objects.select_related("course_class", "course_class__teacher"),
        ),
        pk=session_id,
    )
    if session.status != "open":
        return _error("Buoi diem danh da ket thuc.", status=409)

    try:
        camera_index = int(request.GET.get("camera", 0))
        data = load_encodings(require_data=False)
        webcam = open_webcam(camera_index)
        webcam.set(cv2.CAP_PROP_FRAME_WIDTH, STREAM_WIDTH)
        webcam.set(cv2.CAP_PROP_FRAME_HEIGHT, 360)
        webcam.set(cv2.CAP_PROP_FPS, 20)
        webcam.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    except (TypeError, ValueError) as exc:
        return _error(f"Khong the bat dau camera: {exc}", status=400)
    except Exception as exc:
        return _error(f"Khong the bat dau camera: {exc}", status=503)

    response = StreamingHttpResponse(
        _stream_frames(webcam, session, data),
        content_type="multipart/x-mixed-replace; boundary=frame",
    )
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


@csrf_exempt
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(["GET", "POST"])
def session_list_create(request):
    if request.method == "GET":
        sessions = _scope_sessions(
            request,
            AttendanceSession.objects.select_related(
                "course_class",
                "course_class__teacher",
                "schedule",
                "created_by",
            ),
        )
        return _ok({"results": [_session_payload(session) for session in sessions]})

    try:
        data = _json_body(request)
        user = _default_user(request)
        if not user:
            return _error("Can co it nhat 1 User de gan created_by.", status=400)

        course_class = _check_course_class_access(
            request,
            get_object_or_404(CourseClass, pk=data.get("course_class_id")),
        )
        schedule = None
        if data.get("schedule_id"):
            schedule = get_object_or_404(Schedule, pk=data["schedule_id"])
            _check_course_class_access(request, schedule.course_class)

        session = AttendanceSession.objects.create(
            course_class=course_class,
            schedule=schedule,
            created_by=user,
            status=data.get("status", "open"),
            note=data.get("note", ""),
        )
        return _ok(_session_payload(session), status=201)
    except (ValidationError, IntegrityError) as exc:
        return _error(str(exc), status=400)


@csrf_exempt
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def session_detail(request, pk):
    session = get_object_or_404(
        _scope_sessions(
            request,
            AttendanceSession.objects.select_related(
                "course_class",
                "course_class__teacher",
                "schedule",
                "created_by",
            ),
        ),
        pk=pk,
    )

    if request.method == "GET":
        return _ok(_session_payload(session))

    if request.method == "DELETE":
        session.delete()
        return _ok({"deleted": True})

    try:
        data = _json_body(request)
        absent_created = 0
        with transaction.atomic():
            session = AttendanceSession.objects.select_for_update().select_related(
                "course_class",
                "course_class__teacher",
                "schedule",
                "created_by",
            ).get(pk=session.pk)
            _check_session_access(request, session)

            if "course_class_id" in data:
                session.course_class = _check_course_class_access(
                    request,
                    get_object_or_404(CourseClass, pk=data["course_class_id"]),
                )
            if "schedule_id" in data:
                session.schedule = get_object_or_404(Schedule, pk=data["schedule_id"]) if data["schedule_id"] else None
                if session.schedule:
                    _check_course_class_access(request, session.schedule.course_class)
            if "status" in data:
                if data["status"] == "closed" and not data.get("confirm_close"):
                    return _error("Can xac nhan truoc khi ket thuc buoi diem danh.", status=400)
                session.status = data["status"]
                if data["status"] == "closed" and not session.ended_at:
                    session.ended_at = timezone.now()
            if "note" in data:
                session.note = data["note"]
            session.full_clean()
            session.save()

            if session.status == "closed":
                absent_created = _mark_unrecorded_students_absent(session)

        payload = _session_payload(session)
        payload["absent_created"] = absent_created
        return _ok(payload)
    except (ValidationError, IntegrityError) as exc:
        return _error(str(exc), status=400)


@csrf_exempt
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(["GET", "POST"])
def record_list_create(request):
    if request.method == "GET":
        records = AttendanceRecord.objects.select_related(
            "session",
            "session__course_class",
            "session__course_class__teacher",
            "student",
        )
        if _is_teacher_only(request):
            records = records.filter(session__course_class__teacher=_teacher_for_request(request))
        session_id = request.GET.get("session_id")
        if session_id:
            records = records.filter(session_id=session_id)
        return _ok({"results": [_record_payload(record) for record in records]})

    try:
        data = _json_body(request)
        session = _check_session_access(
            request,
            get_object_or_404(
                AttendanceSession.objects.select_related("course_class", "course_class__teacher"),
                pk=data.get("session_id"),
            ),
        )
        student = get_object_or_404(Student, pk=data.get("student_id"))
        record = AttendanceRecord.objects.create(
            session=session,
            student=student,
            status=data.get("status", "present"),
            method=data.get("method", "manual"),
            confidence=float(data.get("confidence", 0.0)),
            timestamp=timezone.now(),
            note=data.get("note", ""),
        )
        return _ok(_record_payload(record), status=201)
    except (ValidationError, IntegrityError, ValueError) as exc:
        return _error(str(exc), status=400)


@csrf_exempt
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def record_detail(request, pk):
    record = get_object_or_404(
        AttendanceRecord.objects.select_related(
            "session",
            "session__course_class",
            "session__course_class__teacher",
            "student",
        ),
        pk=pk,
    )
    _check_session_access(request, record.session)

    if request.method == "GET":
        return _ok(_record_payload(record))

    if request.method == "DELETE":
        record.delete()
        return _ok({"deleted": True})

    try:
        data = _json_body(request)
        if "session_id" in data:
            record.session = _check_session_access(
                request,
                get_object_or_404(
                    AttendanceSession.objects.select_related("course_class", "course_class__teacher"),
                    pk=data["session_id"],
                ),
            )
        if "student_id" in data:
            record.student = get_object_or_404(Student, pk=data["student_id"])
        for field in ["status", "method", "confidence", "note"]:
            if field in data:
                setattr(record, field, data[field])
        if "timestamp" not in data:
            record.timestamp = timezone.now()
        record.full_clean()
        record.save()
        return _ok(_record_payload(record))
    except (ValidationError, IntegrityError, ValueError) as exc:
        return _error(str(exc), status=400)


@csrf_exempt
@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(["POST"])
def recognize_attendance(request):
    session_id = request.POST.get("session_id")
    image_file = request.FILES.get("image")

    if not session_id:
        return _error("Thieu session_id.", status=400)
    if not image_file:
        return _error("Thieu file anh voi field name la image.", status=400)

    session = _check_session_access(
        request,
        get_object_or_404(
            AttendanceSession.objects.select_related("course_class", "course_class__teacher"),
            pk=session_id,
        ),
    )
    if session.status != "open":
        return _error("Buoi diem danh da ket thuc.", status=409)

    try:
        result = recognize_from_image(image_file)
    except Exception as exc:
        return _error(f"Loi nhan dien khuon mat: {exc}", status=400)

    if not result:
        return _error("Khong nhan dien duoc sinh vien trong anh.", status=404)

    student = get_object_or_404(Student, student_id=result["student_id"])
    is_enrolled = Enrollment.objects.filter(
        course_class=session.course_class,
        student=student,
        is_active=True,
    ).exists()
    if not is_enrolled:
        return _error("Sinh vien nhan dien duoc khong thuoc lop hoc phan nay.", status=400)

    record, created, already_present = _record_face_attendance(
        session,
        student,
        result["confidence"],
        "Auto recognized by face_matcher",
    )
    if not record:
        return _error("Buoi diem danh da ket thuc.", status=409)

    return _ok(
        {
            "created": created,
            "already_present": already_present,
            "match": result,
            "record": _record_payload(record),
        },
        status=201 if created else 200,
    )


@group_required(ADMIN_GROUP_NAME, TEACHER_GROUP_NAME)
@require_http_methods(["GET"])
def export_session_excel(request, session_id):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    session = get_object_or_404(
        _scope_sessions(
            request,
            AttendanceSession.objects.select_related(
                "course_class",
                "course_class__course",
                "course_class__semester",
                "course_class__teacher",
                "course_class__teacher__user",
                "created_by",
            ),
        ),
        pk=session_id,
    )

    course_class = session.course_class
    enrollments = (
        Enrollment.objects.filter(course_class=course_class)
        .select_related("student", "student__student_class", "student__student_class__department")
        .order_by("student__student_id")
    )
    all_records = AttendanceRecord.objects.filter(session__course_class=course_class)
    sessions = AttendanceSession.objects.filter(course_class=course_class).order_by("started_at")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách sinh viên"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A6B3C", end_color="1A6B3C", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"DANH SÁCH SINH VIÊN - LỚP {course_class.class_code}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = align_center

    ws.merge_cells("A2:H2")
    teacher_name = course_class.teacher.user.get_full_name() if course_class.teacher and course_class.teacher.user else ""
    subtitle_cell = ws.cell(row=2, column=1, value=f"Môn học: {course_class.course.course_name} | Giảng viên: {teacher_name}")
    subtitle_cell.font = Font(italic=True)
    subtitle_cell.alignment = align_center

    headers = ["STT", "MSSV", "Họ Tên", "Lớp Sinh Hoạt", "Ngành", "Trạng Thái", "Ngày Đăng Ký", "Tỉ lệ đi học"]
    for attendance_session in sessions:
        headers.append(timezone.localtime(attendance_session.started_at).strftime("%d/%m/%Y"))

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    session_list = list(sessions)
    closed_session_ids = {attendance_session.id for attendance_session in session_list if attendance_session.status == "closed"}

    for row_num, enrollment in enumerate(enrollments, 1):
        student_records = all_records.filter(student=enrollment.student)
        student_records_dict = {record.session_id: record.status for record in student_records}
        s_total = len(set(student_records_dict) | closed_session_ids)
        if s_total > 0:
            s_present = sum(1 for status in student_records_dict.values() if status == "present")
            attendance_rate = f"{round((s_present / s_total) * 100, 1)}%"
        else:
            attendance_rate = "Chưa có DL"

        student_class = enrollment.student.student_class
        row = [
            row_num,
            enrollment.student.student_id,
            enrollment.student.full_name,
            student_class.class_code if student_class else "",
            student_class.department.name if student_class and student_class.department else "",
            "Đang học" if enrollment.is_active else "Nghỉ học",
            timezone.localtime(enrollment.enrolled_at).strftime("%d/%m/%Y"),
            attendance_rate,
        ]

        for attendance_session in session_list:
            status = student_records_dict.get(attendance_session.id)
            if status == "present":
                row.append("Có mặt")
            elif status == "absent":
                row.append("Vắng")
            elif status == "late":
                row.append("Đi trễ")
            elif status == "excused":
                row.append("Có phép")
            elif attendance_session.status == "closed":
                row.append("Vắng")
            else:
                row.append("-")

        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num + 4, column=col_num, value=cell_value)
            cell.border = border
            if col_num in [1, 2, 6, 7, 8] or col_num > 8:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            if cell_value == "Có mặt":
                cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
                cell.font = Font(color="166534")
            elif cell_value == "Vắng":
                cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                cell.font = Font(color="991b1b")
            elif cell_value == "Đi trễ":
                cell.fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
                cell.font = Font(color="854d0e")
            elif cell_value == "Có phép":
                cell.fill = PatternFill(start_color="e0e7ff", end_color="e0e7ff", fill_type="solid")
                cell.font = Font(color="3730a3")

    column_widths = {"A": 6, "B": 15, "C": 30, "D": 15, "E": 25, "F": 15, "G": 15, "H": 12}
    for idx in range(sessions.count()):
        col_letter = openpyxl.utils.get_column_letter(idx + 9)
        column_widths[col_letter] = 15

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="DSSV_{course_class.class_code}.xlsx"'
    wb.save(response)
    return response
